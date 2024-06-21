from ctypes import alignment
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Get the directory of the currently executing script 
script_dir = os.path.dirname(os.path.abspath(__file__))

# Construct the path to the file you want to read
# This makes the assumption that the file is in the same directory as the script
file_name = "RBG UPS Assessment Data.csv"
file_path = os.path.join(script_dir, file_name)
df = pd.read_csv(file_path)

# Sort the DataFrame by 'UPS Score'
df = df.sort_values(by='UPS Score')

columns_to_remove = ['Sku number', 'Manufacturer', 'Score rating', 'External Battery Sku number', 'Output Load (avg 24h in %)',
                      'Load (min 30 days in %)']
df = df.drop(columns=columns_to_remove)

# Save the sorted DataFrame to an Excel file
excel_file_name = "RBG_UPS_Assessment_Data.xlsx"
excel_file_path = os.path.join(script_dir, excel_file_name)
df.to_excel(excel_file_path, index=False, engine='openpyxl')

# Load the Excel file for formatting
wb = load_workbook(excel_file_path)
ws = wb.active

# Define the fills and border
red_fill = PatternFill(start_color="FF3232", end_color="FF3232", fill_type='solid')
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type='solid')
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type='solid')
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type='solid')

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))


#for line in df.columns:
#    print(line)

#legend_font = Font(size = 14)
# Insert legend rows at the top
#ws.insert_rows(1, 5)
#ws['A1'] = 'Legend:'
#ws['A2'] = '0 < UPS Score < 34'
#ws['A3'] = '34 <= UPS Score < 67'
#ws['A4'] = '67 <= UPS Score <= 100'
#ws['A5'] = 'load > 80%'

# Apply fills to the legend cells
#ws['B2'].fill = red_fill
#ws['B3'].fill = yellow_fill
#ws['B4'].fill = green_fill
#ws['B5'].fill = orange_fill

# Align the legend text
#for cell in ['A1', 'A2', 'A3', 'A4', 'A5']:
    #ws[cell].alignment = Alignment(horizontal='left', vertical='center')
    #ws[cell].font = legend_font

# Apply the fills to cells in 'UPS Score' column based on the value
# If the ups score is less than 34, the cell is filled as red.
# If the ups score is less than 67, it is filled in as yellow.
# If the ups score is 67 or greater, it is filled in as green.
ups_score_col_idx = df.columns.get_loc('UPS Score') + 1
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ups_score_col_idx, max_col=ups_score_col_idx):
    for cell in row:
        try:
            cell_value = int(cell.value)
            if cell_value >= 67:
                cell.fill = green_fill
                cell.border = thin_border
            elif cell_value >= 34:
                cell.fill = yellow_fill
                cell.border = thin_border
            elif cell_value > 0:
                cell.fill = red_fill
                cell.border = thin_border
        except (ValueError, TypeError):
            pass  

# Apply the fills to cells in 'Battery Temparture (avg 24h deg C)' column based on value
# If the temperature is greater than 25, but less than 30, the cell is filled as yellow.
# If the temperature is greater than 30, the cell is filled as red.
bat_temp_24_col_idx = df.columns.get_loc('Battery Temperature (avg 24h deg C)') + 1
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=bat_temp_24_col_idx, max_col=bat_temp_24_col_idx):
    for cell in row:
        try:
            
            cell.value = round(cell.value,2)
            cell_value = cell.value
            if cell_value >= 32:
                cell.fill = red_fill
                cell.border = thin_border
            elif cell_value > 25:
                cell.fill = yellow_fill
                cell.border = thin_border
                
        except (ValueError, TypeError):
            pass


# Apply the fills to cells in 'Load (average 30 days in %)' column based on value
# If the load is greater than 80% but less than 100%, the cell is filled as yellow.
# If the load is greater than 100%, the cell is filled in as red.
load_avg_30_col_idx = df.columns.get_loc('Load (average 30 days in %)') + 1
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=load_avg_30_col_idx, max_col=load_avg_30_col_idx):
    for cell in row:
        try:
            
            cell.value = round(cell.value,2)
            cell_value = cell.value
            if cell_value >= 100:
                cell.fill = red_fill
                cell.border = thin_border
            elif cell_value >= 80:
                cell.fill = yellow_fill
                cell.border = thin_border
                
        except (ValueError, TypeError):
            pass

# Apply the fills to cells in 'Load (max 30 days in %)' column based on value
# If the load is greater than 80% but less than 100%, the cell is filled as yellow.
# If the load is greater than 100%, the cell is filled in as red.
load_max_30_col_idx = df.columns.get_loc('Load (max 30 days in %)') + 1
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=load_max_30_col_idx, max_col=load_max_30_col_idx):
    for cell in row:
        try:
            
            cell.value = round(cell.value,2)
            cell_value = cell.value
            if cell_value >= 100:
                cell.fill = red_fill
                cell.border = thin_border
            elif cell_value >= 80:
                cell.fill = yellow_fill
                cell.border = thin_border
            
        except (ValueError, TypeError):
            pass 

        

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column letter
    for cell in col:
        try:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2  # Add a little extra space
    ws.column_dimensions[column].width = adjusted_width

# Auto-adjust row heights
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    max_height = 15  # Set a default minimum height
    for cell in row:
        if cell.value is not None:
            # Calculate the height needed for the cell's content
            lines = str(cell.value).split('\n')
            max_height = max(max_height, len(lines) * 15) 
    ws.row_dimensions[row[0].row].height = max_height

# Save the modified Excel file
wb.save(excel_file_path)

# Automatically open the Excel file
os.system(f"open '{excel_file_path}'")
